import openpyxl, json

wb = openpyxl.load_workbook('PATH_TO_YOUR_SILVER_FOX_XLSX')
print('工作表:', wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== 工作表: {name} ({ws.max_row}行 x {ws.max_column}列) ===')
    
    # 打印表头
    headers = [c.value for c in ws[1]]
    print('表头:', headers)
    
    # 打印前10行
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(11, ws.max_row), values_only=True), 1):
        vals = [str(v)[:30] if v else '' for v in row]
        print(f'  {i}: ' + ' | '.join(vals))
    
    # 如果有更多行，统计
    if ws.max_row > 11:
        print(f'  ... 还有 {ws.max_row - 11} 行')

# 保存完整内容到json
data = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v else '' for v in row])
    data[name] = rows

with open('E:/Reasonix/.reasonix/skills/ir-forensic-analysis/silver_fox_xlsx_data.json', 'w') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print(f'\n完整数据已保存')
