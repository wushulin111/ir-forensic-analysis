#!/usr/bin/env python3
"""解析银狐情报Excel并更新到规则库"""
import openpyxl, json, re, os

XLSX = 'PATH_TO_YOUR_SILVER_FOX_XLSX'
RULES_DIR = 'E:/Reasonix/.reasonix/skills/ir-forensic-analysis/rules'
wb = openpyxl.load_workbook(XLSX)

# ========== 1. 提取恶意域名/IP（情报总表） ==========
ws = wb['情报总表']
domains = []
ips = []
hashes_from_total = []

for row in ws.iter_rows(min_row=3, values_only=True):
    seq = row[0]  # 序号
    fake_software = str(row[1] or '')
    domain_raw = str(row[2] or '')
    hash_val = str(row[8] or '').strip()
    
    # 提取域名（去掉[.]符号）
    d = domain_raw.replace('[.]', '.').replace('（恶意域名','').strip()
    if d and d != 'None':
        # 判断是IP还是域名
        ip_match = re.match(r'^\d+\.\d+\.\d+\.\d+', d)
        if ip_match:
            ips.append(ip_match.group())
        elif '.' in d:
            domains.append(d)
    
    # 提取hash
    hash_match = re.search(r'[a-fA-F0-9]{32}', hash_val)
    if hash_match:
        hashes_from_total.append(hash_match.group())

print(f'情报总表: 域名 {len(set(domains))} 个, IP {len(set(ips))} 个, Hash {len(set(hashes_from_total))} 个')

# ========== 2. 提取释放路径 ==========
ws2 = wb['释放路径']
paths = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    p = str(row[0] or '').strip()
    if p and p != 'None':
        paths.append(p)

print(f'释放路径: {len(set(paths))} 条')

# ========== 3. 提取样本哈希 ==========
ws3 = wb['样本哈希值']
hashes = []
for row in ws3.iter_rows(min_row=2, values_only=True):
    h = str(row[0] or '').strip().lower()
    if h and h != 'None' and re.match(r'^[a-f0-9]{30,40}$', h):
        hashes.append(h)

print(f'样本哈希: {len(set(hashes))} 个 (去重后)')

# ========== 4. 提取银狐特征 ==========
ws4 = wb['银狐特征']
features = {'file_name_patterns': [], 'file_paths': [], 'hashes_feature': []}
for row in ws4.iter_rows(min_row=2, values_only=True):
    if row[1] and str(row[1]).strip() != 'None':
        features['file_name_patterns'].append(str(row[1]).strip())
    if row[2] and str(row[2]).strip() != 'None':
        features['file_paths'].append(str(row[2]).strip())
    if row[3] and str(row[3]).strip() != 'None':
        h = str(row[3]).strip().lower()
        if re.match(r'^[a-f0-9]{30,40}$', h):
            features['hashes_feature'].append(h)

print(f'银狐特征: 文件名模式 {len(features["file_name_patterns"])} 条, 路径 {len(features["file_paths"])} 条')

# ========== 5. 更新 IOC 库 ==========
all_hashes = list(set(hashes + hashes_from_total + features['hashes_feature']))
all_domains = list(set(domains))
all_ips = list(set(ips))
all_paths = list(set(paths + features['file_paths']))

ioc_lib = {
    "description": "银狐木马IOC库 - 来自开源恶意域名情报库",
    "last_updated": "2026-06-01",
    "source": "开源恶意域名情报库【挖矿 勒索 银狐】(腾讯文档)",
    "stats": {
        "domains": len(all_domains),
        "ips": len(all_ips),
        "hashes": len(all_hashes),
        "file_paths": len(all_paths)
    },
    "iocs": {
        "domains": sorted(all_domains)[:500],  # 限制500条避免文件过大
        "ips": sorted(all_ips),
        "hashes": sorted(all_hashes),
        "file_paths": sorted(all_paths)[:200]
    }
}

# 保存
with open(os.path.join(RULES_DIR, 'ioc_library.json'), 'w', encoding='utf-8') as f:
    json.dump(ioc_lib, f, ensure_ascii=False, indent=2)

print(f'\n已保存到 ioc_library.json')
print(f'  - 域名: {len(all_domains)} 个')
print(f'  - IP: {len(all_ips)} 个')
print(f'  - Hash: {len(all_hashes)} 个')
print(f'  - 释放路径: {len(all_paths)} 条')

# ========== 6. 添加到 silver_fox_rules.json ==========
with open(os.path.join(RULES_DIR, 'silver_fox_rules.json'), 'r', encoding='utf-8') as f:
    rules = json.load(f)

# 更新 SFOX-0023 为域名规则，加入实际域名
new_domain_rule = {
    "rule_id": "SFOX-0024",
    "type": "ioc_domain",
    "name": "银狐恶意域名库（开源情报）",
    "description": f"来自开源恶意域名情报库的银狐仿冒/钓鱼域名，共{len(all_domains)}个",
    "tag": "fox_tag",
    "domains": all_domains[:200],  # 限制200条
    "source": "开源恶意域名情报库 2026-06-01",
    "severity": "critical",
    "confidence": 0.9,
    "mitre_attack": "T1071",
    "created_at": "2026-06-01",
    "hit_count": 0
}

new_hash_rule = {
    "rule_id": "SFOX-0025",
    "type": "ioc_hash",
    "name": "银狐样本哈希库（开源情报）",
    "description": f"来自开源恶意域名情报库的银狐样本哈希，共{len(all_hashes)}个",
    "tag": "fox_tag",
    "hashes": all_hashes[:200],  # 限制200条
    "source": "开源恶意域名情报库 2026-06-01",
    "severity": "critical",
    "confidence": 0.9,
    "mitre_attack": "T1204.002",
    "created_at": "2026-06-01",
    "hit_count": 0
}

new_path_rule = {
    "rule_id": "SFOX-0026",
    "type": "ioc_filepath",
    "name": "银狐释放路径库（开源情报）",
    "description": f"来自开源恶意域名情报库的银狐样本释放路径，共{len(all_paths)}条",
    "tag": "fox_tag",
    "file_paths": all_paths[:200],  # 限制200条
    "source": "开源恶意域名情报库 2026-06-01",
    "severity": "high",
    "confidence": 0.85,
    "mitre_attack": "T1204.002",
    "created_at": "2026-06-01",
    "hit_count": 0
}

rules.append(new_domain_rule)
rules.append(new_hash_rule)
rules.append(new_path_rule)

with open(os.path.join(RULES_DIR, 'silver_fox_rules.json'), 'w', encoding='utf-8') as f:
    json.dump(rules, f, ensure_ascii=False, indent=2)

print(f'\n已更新 silver_fox_rules.json (共 {len(rules)} 条规则)')
print(f'  新增: SFOX-0024 域名库, SFOX-0025 哈希库, SFOX-0026 释放路径库')
